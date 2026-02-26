"""Shared Excel styling constants."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelStyles:
    """Centralized styling for the Excel report."""

    # Header
    HEADER_FILL = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    HEADER_FONT = Font(
        name="B Nazanin", bold=True, color="FFFFFF", size=11
    )
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    # Status fills
    CONFIRM_FILL = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    REVIEW_FILL = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    REJECT_FILL = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )

    # General
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    BODY_FONT = Font(name="B Nazanin", size=11)
    TITLE_FONT = Font(name="B Nazanin", bold=True, size=16, color="1F4E79")
    SECTION_FONT = Font(name="B Nazanin", bold=True, size=13, color="1F4E79")
    SUCCESS_FONT = Font(name="B Nazanin", size=14, color="00B050")
