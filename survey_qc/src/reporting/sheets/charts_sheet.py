"""Charts sheet with 6 visualizations."""

from openpyxl.chart import BarChart, PieChart, Reference

from src.models.entities import QCResult
from src.models.enums import QCStatus

from .base_sheet import BaseSheetWriter


class ChartsSheetWriter(BaseSheetWriter):

    def __init__(self, workbook, results: list[QCResult]):
        super().__init__(workbook)
        self._results = results

    def write(self):
        ws = self._create_sheet("📈 نمودارها")

        next_data_row = 1
        chart_col = "D"
        chart_top_row = 1

        next_data_row = self._chart_status_pie(ws, next_data_row, chart_col, chart_top_row)
        chart_top_row += 17
        next_data_row = self._chart_contradiction_types(ws, next_data_row + 2, chart_col, chart_top_row)
        chart_top_row += 17
        next_data_row = self._chart_score_distribution(ws, next_data_row + 2, chart_col, chart_top_row)
        chart_top_row += 17
        next_data_row = self._chart_reliability_distribution(ws, next_data_row + 2, chart_col, chart_top_row)
        chart_top_row += 17
        next_data_row = self._chart_delivery_time(ws, next_data_row + 2, chart_col, chart_top_row)
        chart_top_row += 17
        self._chart_agency_contradictions(ws, next_data_row + 2, chart_col, chart_top_row)

        return ws

    def _chart_status_pie(self, ws, start_row, chart_col, chart_top_row) -> int:
        confirmed = sum(1 for r in self._results if r.status == QCStatus.CONFIRM)
        review = sum(1 for r in self._results if r.status == QCStatus.REVIEW)
        rejected = sum(1 for r in self._results if r.status == QCStatus.REJECT)

        ws.cell(row=start_row, column=1, value="وضعیت")
        ws.cell(row=start_row, column=2, value="تعداد")
        data = [("تأیید", confirmed), ("بررسی", review), ("رد", rejected)]
        for i, (label, val) in enumerate(data, start=start_row + 1):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=val)

        pie = PieChart()
        pie.title = "توزیع وضعیت نظرسنجی‌ها"
        pie.style = 10
        pie.width, pie.height = 18, 12
        labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + 3)
        chart_data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + 3)
        pie.add_data(chart_data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, f"{chart_col}{chart_top_row}")

        return start_row + 4

    def _chart_contradiction_types(self, ws, start_row, chart_col, chart_top_row) -> int:
        type_counts = {}
        for r in self._results:
            for c in r.contradictions:
                key = c.contradiction_type.value
                type_counts[key] = type_counts.get(key, 0) + 1

        if not type_counts:
            return start_row

        ws.cell(row=start_row, column=1, value="نوع تناقض")
        ws.cell(row=start_row, column=2, value="تعداد")
        for i, (ctype, count) in enumerate(type_counts.items(), start=start_row + 1):
            ws.cell(row=i, column=1, value=ctype)
            ws.cell(row=i, column=2, value=count)

        bar = BarChart()
        bar.title = "تفکیک تناقضات بر اساس نوع"
        bar.style = 10
        bar.y_axis.title = "تعداد"
        bar.width, bar.height = 18, 12
        bar_data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(type_counts))
        bar_cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(type_counts))
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_cats)
        ws.add_chart(bar, f"{chart_col}{chart_top_row}")

        return start_row + len(type_counts) + 1

    def _chart_score_distribution(self, ws, start_row, chart_col, chart_top_row) -> int:
        score_counts = {}
        for r in self._results:
            s = int(r.score) if r.score == int(r.score) else r.score
            score_counts[s] = score_counts.get(s, 0) + 1

        if not score_counts:
            return start_row

        ws.cell(row=start_row, column=1, value="امتیاز")
        ws.cell(row=start_row, column=2, value="تعداد")
        sorted_scores = sorted(score_counts.keys())
        for i, score_val in enumerate(sorted_scores, start=start_row + 1):
            ws.cell(row=i, column=1, value=score_val)
            ws.cell(row=i, column=2, value=score_counts[score_val])

        bar = BarChart()
        bar.title = "توزیع امتیاز مشتریان"
        bar.style = 10
        bar.y_axis.title = "تعداد"
        bar.width, bar.height = 18, 12
        bar_data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(sorted_scores))
        bar_cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(sorted_scores))
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_cats)
        ws.add_chart(bar, f"{chart_col}{chart_top_row}")

        return start_row + len(sorted_scores) + 1

    def _chart_reliability_distribution(self, ws, start_row, chart_col, chart_top_row) -> int:
        bins = [("0-20", 0, 20), ("20-40", 20, 40), ("40-60", 40, 60),
                ("60-80", 60, 80), ("80-100", 80, 101)]

        ws.cell(row=start_row, column=1, value="بازه اعتبار")
        ws.cell(row=start_row, column=2, value="تعداد")
        for i, (label, low, high) in enumerate(bins, start=start_row + 1):
            count = sum(1 for r in self._results if low <= r.reliability_score < high)
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=count)

        bar = BarChart()
        bar.title = "توزیع امتیاز اعتبار"
        bar.style = 10
        bar.y_axis.title = "تعداد"
        bar.width, bar.height = 18, 12
        bar_data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(bins))
        bar_cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(bins))
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_cats)
        ws.add_chart(bar, f"{chart_col}{chart_top_row}")

        return start_row + len(bins) + 1

    def _chart_delivery_time(self, ws, start_row, chart_col, chart_top_row) -> int:
        delivery_times = [r.delivery_hours for r in self._results if r.delivery_hours is not None]
        if not delivery_times:
            return start_row

        dt_bins = [("0-6", 0, 6), ("6-12", 6, 12), ("12-24", 12, 24),
                   ("24-48", 24, 48), ("48+", 48, 9999)]

        ws.cell(row=start_row, column=1, value="بازه تحویل (ساعت)")
        ws.cell(row=start_row, column=2, value="تعداد")
        for i, (label, low, high) in enumerate(dt_bins, start=start_row + 1):
            count = sum(1 for h in delivery_times if low <= h < high)
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=count)

        bar = BarChart()
        bar.title = "توزیع زمان تحویل"
        bar.style = 10
        bar.y_axis.title = "تعداد"
        bar.width, bar.height = 18, 12
        bar_data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(dt_bins))
        bar_cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(dt_bins))
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_cats)
        ws.add_chart(bar, f"{chart_col}{chart_top_row}")

        return start_row + len(dt_bins) + 1

    def _chart_agency_contradictions(self, ws, start_row, chart_col, chart_top_row) -> int:
        agency_counts = {}
        for r in self._results:
            agency_counts[r.agency] = agency_counts.get(r.agency, 0) + len(r.contradictions)

        if not agency_counts:
            return start_row

        top = sorted(agency_counts.items(), key=lambda x: x[1], reverse=True)[:10]

        ws.cell(row=start_row, column=1, value="نمایندگی")
        ws.cell(row=start_row, column=2, value="تعداد تناقض")
        for i, (agency, count) in enumerate(top, start=start_row + 1):
            ws.cell(row=i, column=1, value=agency)
            ws.cell(row=i, column=2, value=count)

        bar = BarChart()
        bar.title = "نمایندگی‌ها بر اساس تعداد تناقض"
        bar.style = 10
        bar.width, bar.height = 18, 12
        bar_data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(top))
        bar_cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(top))
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_cats)
        ws.add_chart(bar, f"{chart_col}{chart_top_row}")

        return start_row + len(top) + 1
