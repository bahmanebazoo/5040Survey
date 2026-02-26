"""Summary dashboard sheet."""

import numpy as np
from openpyxl.styles import Alignment, Font

from src.models.entities import QCResult
from src.models.enums import QCStatus

from .base_sheet import BaseSheetWriter


class SummarySheetWriter(BaseSheetWriter):

    def __init__(self, workbook, results: list[QCResult]):
        super().__init__(workbook)
        self._results = results

    def write(self):
        ws = self._create_sheet("📊 خلاصه داشبورد")

        total = len(self._results)
        confirmed = sum(1 for r in self._results if r.status == QCStatus.CONFIRM)
        review = sum(1 for r in self._results if r.status == QCStatus.REVIEW)
        rejected = sum(1 for r in self._results if r.status == QCStatus.REJECT)
        avg_reliability = np.mean([r.reliability_score for r in self._results]) if self._results else 0
        avg_score = np.mean([r.score for r in self._results]) if self._results else 0
        total_contradictions = sum(len(r.contradictions) for r in self._results)

        type_counts = {}
        for r in self._results:
            for c in r.contradictions:
                key = c.contradiction_type.value
                type_counts[key] = type_counts.get(key, 0) + 1

        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "🔍 داشبورد کنترل کیفیت نظرسنجی مشتریان"
        title_cell.font = self._styles.TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center")

        # Metrics
        metrics = [
            ("📋 تعداد کل نظرسنجی‌ها", total),
            ("✅ تأیید شده", f"{confirmed} ({confirmed/total*100:.1f}%)" if total else "0"),
            ("⚠️ نیاز به بررسی", f"{review} ({review/total*100:.1f}%)" if total else "0"),
            ("❌ رد شده", f"{rejected} ({rejected/total*100:.1f}%)" if total else "0"),
            ("📊 میانگین امتیاز اعتبار", f"{avg_reliability:.1f}"),
            ("⭐ میانگین امتیاز مشتری", f"{avg_score:.2f}"),
            ("🔴 تعداد کل تناقضات", total_contradictions),
        ]

        for i, (label, value) in enumerate(metrics, start=3):
            ws.cell(row=i, column=1, value=label).font = Font(
                name="B Nazanin", bold=True, size=12
            )
            ws.cell(row=i, column=3, value=str(value)).font = Font(
                name="B Nazanin", size=12
            )

        # Contradiction breakdown
        row_start = len(metrics) + 5
        ws.cell(row=row_start, column=1, value="📋 تفکیک تناقضات").font = \
            self._styles.SECTION_FONT
        row_start += 1
        for col_idx, header in enumerate(["نوع تناقض", "تعداد"], start=1):
            cell = ws.cell(row=row_start, column=col_idx, value=header)
            cell.font = self._styles.HEADER_FONT
            cell.fill = self._styles.HEADER_FILL

        for i, (ctype, count) in enumerate(type_counts.items(), start=row_start + 1):
            ws.cell(row=i, column=1, value=ctype).font = self._styles.BODY_FONT
            ws.cell(row=i, column=2, value=count).font = self._styles.BODY_FONT

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 25

        return ws
