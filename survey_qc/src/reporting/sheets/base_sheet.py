"""Abstract base for all sheet writers."""

from abc import ABC, abstractmethod

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models.enums import QCStatus
from src.reporting.styles import ExcelStyles


class BaseSheetWriter(ABC):
    """
    Base class for individual sheet writers.
    Provides shared formatting utilities.
    """

    def __init__(self, workbook: Workbook):
        self._wb = workbook
        self._styles = ExcelStyles()

    @abstractmethod
    def write(self) -> Worksheet:
        """Create and populate the sheet. Return the worksheet."""
        ...

    # ---- Shared Utilities ----

    def _create_sheet(self, title: str) -> Worksheet:
        ws = self._wb.create_sheet(title)
        ws.sheet_view.rightToLeft = True
        return ws

    def _write_header_row(self, ws: Worksheet, headers: list[str]):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self._styles.HEADER_FONT
            cell.fill = self._styles.HEADER_FILL
            cell.alignment = self._styles.HEADER_ALIGNMENT
            cell.border = self._styles.BORDER

    def _apply_status_fill(self, cell, status: QCStatus):
        if status == QCStatus.CONFIRM:
            cell.fill = self._styles.CONFIRM_FILL
        elif status == QCStatus.REVIEW:
            cell.fill = self._styles.REVIEW_FILL
        elif status == QCStatus.REJECT:
            cell.fill = self._styles.REJECT_FILL

    @staticmethod
    def _auto_fit_columns(ws: Worksheet, headers: list[str]):
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(header) * 1.5, 14)
